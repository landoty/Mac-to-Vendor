import re, sys
import openpyxl

def mac_to_oui(mac) -> str:
    mac = re.sub(":", "",mac)
    mac = mac.upper()
    return(mac[:6])

def oui_to_vendor(oui) -> str:
    with open("oui_vendors.txt", "r", encoding="utf8") as vendor_file:
        for line in vendor_file:
            if re.search(oui, line):
                return(line[22:])

def main() -> None:
    path = sys.argv[1]
    break_line = "-----------------------------------"
    print(break_line)
    mac_column = int(input("MAC Address Column Number: "))
    print(break_line)

    print("Opening excel sheet...")
    wb = openpyxl.load_workbook(path)
    print(f"Opened!\n{break_line}\nWorking...")

    sheet = wb.active
    sheet.insert_cols(mac_column+1)
    sheet.cell(1,mac_column+1).value="Node Vendor"
    row_count = sheet.max_row

    known_vendors = {}
    for i in range(2,row_count):
        mac = sheet.cell(i,mac_column).value
        oui = mac_to_oui(mac)
        if oui in known_vendors:
            sheet.cell(i,mac_column+1).value=known_vendors[oui]
        else:
            vendor = oui_to_vendor(oui)
            sheet.cell(i,mac_column+1).value=vendor
            known_vendors[oui] = vendor
    wb.save(path)
    print(f"Vendor report complete.\nSee updated {path}\n{break_line}")

if __name__ == "__main__":
    main()
